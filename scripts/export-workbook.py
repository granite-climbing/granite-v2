#!/usr/bin/env python3
"""
export-workbook.py  —  Granite Phase 2 workbook exporter.

Usage:
    python3 scripts/export-workbook.py \
        docs/data/granite-v2-migration-prep.xlsx \
        docs/data/granite-v2-workbook.json
"""

import json
import re
import sys
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Header normalization
# ---------------------------------------------------------------------------

def normalize_header(raw: str) -> str:
    """
    Converts a raw header string to a clean snake_case key.
    Steps:
      1. Strip parenthetical annotations, e.g. "(TO FILL)", "(auto)", "(검토: …)"
      2. Lowercase
      3. Trim and replace runs of non-alphanumeric chars with a single underscore
      4. Strip leading/trailing underscores
    """
    # Remove anything inside parentheses (and the parentheses themselves)
    cleaned = re.sub(r"\([^)]*\)", "", raw)
    # Also handle "[…]" brackets if any
    cleaned = re.sub(r"\[[^\]]*\]", "", cleaned)
    cleaned = cleaned.lower().strip()
    # Replace any run of non-alphanumeric characters with a single underscore
    cleaned = re.sub(r"[^a-z0-9]+", "_", cleaned)
    # Collapse repeated underscores and strip edges
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned


def should_drop_key(key: str) -> bool:
    """
    Returns True for keys that should be dropped from the output:
    - Empty string (e.g. 비고 → '' after normalization, or empty header cell)
    - Anything starting with "v1" (legacy v1 columns)
    - Reference-only columns
    """
    if key == "":
        return True
    if key.startswith("v1"):
        return True
    # "소속_v1_boulders" reference column
    if "v1" in key:
        return True
    # access_desc and parking_desc per import contract
    if key in ("access_desc", "parking_desc"):
        return True
    return False


# ---------------------------------------------------------------------------
# Sheet-specific header aliasing
# ---------------------------------------------------------------------------

def alias_headers(raw_headers: list[str | None], sheet_name: str) -> list[str | None]:
    """
    Returns a parallel list of final key names (or None for dropped columns).
    Handles the English-name / Korean-name collision and drops v1/notes columns.
    """
    result: list[str | None] = []
    for i, raw in enumerate(raw_headers):
        if raw is None or str(raw).strip() == "":
            result.append(None)
            continue

        raw_str = str(raw)

        # Position-based alias: distinguish English name from Korean name.
        # The raw header literally contains "English" → alias to name_en.
        if re.search(r"\bEnglish\b", raw_str, re.IGNORECASE):
            result.append("name_en")
            continue

        # Normalize generically
        key = normalize_header(raw_str)

        # Rename "hashtags_json" (after stripping parens/suffix) → "hashtags"
        if key == "hashtags_json":
            key = "hashtags"

        # Rename "summary_…" variants → "summary"
        if key.startswith("summary"):
            key = "summary"

        # Drop columns per rules above
        if should_drop_key(key):
            result.append(None)
            continue

        result.append(key)

    return result


# ---------------------------------------------------------------------------
# Per-sheet processing
# ---------------------------------------------------------------------------

SHEETS = ["Areas", "Crags", "Sectors", "Boulders", "Topos", "Routes"]

# Primary identity column per sheet (used for blank-row filtering)
IDENTITY_COL: dict[str, str] = {
    "Areas":    "slug",
    "Crags":    "slug",
    "Sectors":  "slug",
    "Boulders": "slug",
    "Topos":    "slug",
    "Routes":   "slug",
}


def process_sheet(ws, sheet_name: str) -> list[dict]:
    raw_headers = [cell.value for cell in ws[1]]
    keys = alias_headers(raw_headers, sheet_name)

    identity_key = IDENTITY_COL[sheet_name]

    # Find the position of the identity column
    try:
        identity_pos = keys.index(identity_key)
    except ValueError:
        raise RuntimeError(
            f"Sheet '{sheet_name}': identity column '{identity_key}' not found in headers. "
            f"Resolved keys: {keys}"
        )

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        identity_val = row[identity_pos]
        if identity_val is None or str(identity_val).strip() == "":
            continue  # blank / trailing empty row

        record: dict = {}
        for col_idx, key in enumerate(keys):
            if key is None:
                continue  # dropped column
            val = row[col_idx] if col_idx < len(row) else None
            # Normalize empty strings to None for cleaner JSON
            if isinstance(val, str) and val.strip() == "":
                val = None
            record[key] = val

        rows.append(record)

    return rows


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <input.xlsx> <output.json>", file=sys.stderr)
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    json_path = Path(sys.argv[2])

    if not xlsx_path.exists():
        print(f"ERROR: xlsx not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    output: dict[str, list] = {}
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: sheet '{sheet_name}' not found in workbook", file=sys.stderr)
            sys.exit(1)
        rows = process_sheet(wb[sheet_name], sheet_name)
        key = sheet_name.lower()
        output[key] = rows
        print(f"{sheet_name}: {len(rows)} rows")

    json_path.parent.mkdir(parents=True, exist_ok=True)
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\nWrote {json_path}")


if __name__ == "__main__":
    main()
